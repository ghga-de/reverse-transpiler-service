# Copyright 2021 - 2025 Universität Tübingen, DKFZ, EMBL, and Universität zu Köln
# for the German Human Genome-Phenome Archive (GHGA)
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""Integration tests for the reverse transpiler."""

from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from rts.models import StudyMetadata
from rts.ports.inbound.rev_tran import ReverseTranspilerPort
from tests.fixtures.joint import JointFixture

pytestmark = pytest.mark.asyncio()


def assert_workbooks_match(expected: Workbook, actual: Workbook) -> None:
    """Compare two workbooks for equality.

    This is more reliable than comparing bytestreams directly.
    """
    assert sorted(expected.sheetnames) == sorted(actual.sheetnames), (
        "Sheet names do not match"
    )

    for sheet_name in expected.sheetnames:
        expected_sheet = expected[sheet_name]
        actual_sheet = actual[sheet_name]

        expected_row_count = len([_ for _ in expected_sheet.rows])
        actual_row_count = len([_ for _ in actual_sheet.rows])
        assert expected_row_count == actual_row_count, (
            f"Row counts do not match for sheet: {sheet_name} (expected:"
            + f" {expected_row_count}, got: {actual_row_count})"
        )

        for row_index, row in enumerate(expected_sheet.iter_rows(), 1):
            for col_index, expected_cell in enumerate(row, 1):
                actual_value = actual_sheet.cell(row=row_index, column=col_index).value
                assert expected_cell.value == actual_value, (
                    f"Cell values do not match at {sheet_name} row {row_index},"
                    + f" column {col_index}. Expected: {expected_cell.value},"
                    + f" Got: {actual_value}"
                )


async def load_test_data(reverse_transpiler: ReverseTranspilerPort) -> StudyMetadata:
    """Load the test data by calling `reverse_transpiler.upsert_metadata()`

    Returns the test data as a `StudyMetadata` object.
    """
    with open("tests/fixtures/test_artifact.json") as file:
        study_metadata_json = file.read()
    study_metadata = StudyMetadata.model_validate_json(study_metadata_json)
    await reverse_transpiler.upsert_metadata(study_metadata=study_metadata)
    return study_metadata


async def test_basic_reverse_transpilation(joint_fixture: JointFixture):
    """Test basic reverse transpilation functionality."""
    # Load a sample study metadata JSON file into the database
    reverse_transpiler = joint_fixture.reverse_transpiler
    study_metadata = await load_test_data(reverse_transpiler)

    # Run the reverse transpilation manually so we have the output on hand to compare
    expected = reverse_transpiler._reverse_transpile(study_metadata)  # type: ignore
    assert isinstance(expected, Workbook)

    # Call the retrieval function
    retrieved_bytes = await reverse_transpiler.retrieve_workbook(
        study_accession=study_metadata.study_accession
    )
    assert isinstance(retrieved_bytes, bytes)

    # Load the retrieved workbook and compare it to the expected one.
    retrieved = load_workbook(BytesIO(retrieved_bytes))
    assert isinstance(retrieved, Workbook)
    assert_workbooks_match(expected=expected, actual=retrieved)


async def test_rest(joint_fixture: JointFixture):
    """Test that the spreadsheet data is delivered via REST API intact."""
    # Load a sample study metadata JSON file into the database
    reverse_transpiler = joint_fixture.reverse_transpiler
    study_metadata = await load_test_data(reverse_transpiler)
    accession = study_metadata.study_accession

    # Run the reverse transpilation manually so we have the output on hand to compare
    expected = reverse_transpiler._reverse_transpile(study_metadata)  # type: ignore

    # Now call the REST API to get the transpiled data as well
    response = await joint_fixture.rest_client.get(f"/studies/{accession}")
    assert response.status_code == 200
    assert isinstance(response.content, bytes)

    assert response.headers["Content-Disposition"] == (
        f'attachment; filename="{accession}.xlsx"'
    )
    content_type = response.headers["Content-Type"]
    expected_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert content_type == expected_type

    # Check if the HTTP response content matches the expected workbook
    received = load_workbook(BytesIO(response.content))
    assert_workbooks_match(expected, received)


async def test_upsert(joint_fixture: JointFixture):
    """Test to ensure the core updates the stored workbook data when it consumes
    an updated artifact, and that it doesn't change anything when the data is identical.
    """
    # Load a sample study metadata JSON file into the database
    reverse_transpiler = joint_fixture.reverse_transpiler
    study_metadata = await load_test_data(reverse_transpiler)
    original = reverse_transpiler._reverse_transpile(study_metadata)  # type: ignore
    accession = study_metadata.study_accession

    # Delete the workbook from storage, since that is only created when new data comes in
    await reverse_transpiler._workbook_dao.delete(id_=accession)  # type: ignore

    # Load the data again, but this time it shouldn't recreate the workbook
    await reverse_transpiler.upsert_metadata(study_metadata=study_metadata)
    with pytest.raises(reverse_transpiler.MetadataNotFoundError):
        await reverse_transpiler.retrieve_workbook(study_accession=accession)

    # Change the study metadata now because we want to trigger an upsert
    study_metadata.content["samples"][0]["accession"] = "updated_sample_accession"
    expected_updated = reverse_transpiler._reverse_transpile(study_metadata)  # type: ignore

    # Load the changed data to trigger re-processing and get the workbook
    await reverse_transpiler.upsert_metadata(study_metadata=study_metadata)
    updated_retrieved_bytes = await reverse_transpiler.retrieve_workbook(
        study_accession=accession
    )
    updated_retrieved = load_workbook(BytesIO(updated_retrieved_bytes))

    # Make sure the workbooks match after the update
    assert_workbooks_match(expected=expected_updated, actual=updated_retrieved)

    # Make sure the new workbook is different from the old one
    with pytest.raises(AssertionError):
        assert_workbooks_match(expected=original, actual=updated_retrieved)


async def test_data_deletion(joint_fixture: JointFixture):
    """Test that the metadata and workbook data can be deleted properly"""
    # Load a sample study metadata JSON file into the database
    reverse_transpiler = joint_fixture.reverse_transpiler
    study_metadata = await load_test_data(reverse_transpiler)

    # Delete the metadata
    accession = study_metadata.study_accession
    await reverse_transpiler.delete_metadata(study_accession=accession)

    # Verify it's gone
    with pytest.raises(reverse_transpiler.MetadataNotFoundError):
        await reverse_transpiler.retrieve_metadata(study_accession=accession)
    with pytest.raises(reverse_transpiler.MetadataNotFoundError):
        await reverse_transpiler.retrieve_workbook(study_accession=accession)


async def test_delete_non_existent_metadata(joint_fixture: JointFixture):
    """Test that deleting non-existent metadata does not raise an error."""
    reverse_transpiler = joint_fixture.reverse_transpiler

    # This should not raise an error, even though the metadata does not exist
    await reverse_transpiler.delete_metadata(study_accession="non_existent_accession")


async def test_event_handling(joint_fixture: JointFixture):
    """Test that the reverse transpiler can handle upsertion/deletion events.

    This is a semi-redundant test of the what exists in test_event_sub.py.
    """
    accession = "test_accession"

    # Publish the upsert event
    await joint_fixture.kafka.publish_event(
        topic=joint_fixture.config.artifact_topic,
        type_="upserted",
        key="added_accessions:test_accession",
        payload={
            "artifact_name": "added_accessions",
            "content": {
                "samples": [{"accession": "sample1"}],
                "studies": [{"accession": accession}],
            },
            "study_accession": accession,
        },
    )

    # Verify that nothing is in the DB yet
    with pytest.raises(joint_fixture.reverse_transpiler.MetadataNotFoundError):
        await joint_fixture.reverse_transpiler.retrieve_workbook(
            study_accession=accession
        )

    # Now run the event subscriber to process the upsert event
    await joint_fixture.event_subscriber.run(forever=False)

    # Verify that the workbook is now in the DB
    retrieved_workbook_bytes = await joint_fixture.reverse_transpiler.retrieve_workbook(
        study_accession=accession
    )

    # Cursory check
    assert isinstance(retrieved_workbook_bytes, bytes)
    retrieved_workbook = load_workbook(BytesIO(retrieved_workbook_bytes))
    assert retrieved_workbook.sheetnames == ["Sample", "Study"]  # from config

    # Publish a deletion event
    await joint_fixture.kafka.publish_event(
        topic=joint_fixture.config.artifact_topic,
        type_="deleted",
        key=f"added_accessions:{accession}",
        payload={},
    )

    # Run the event subscriber to process the deletion event
    await joint_fixture.event_subscriber.run(forever=False)

    # Verify that the workbook is no longer in the DB
    with pytest.raises(joint_fixture.reverse_transpiler.MetadataNotFoundError):
        await joint_fixture.reverse_transpiler.retrieve_workbook(
            study_accession=accession
        )
