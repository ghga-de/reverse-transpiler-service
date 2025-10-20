# Copyright 2021 - 2025 Universität Tübingen, DKFZ, EMBL, and Universität zu Köln
# for the German Human Genome-Phenome Archive (GHGA)
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""Core functionality of the reverse transpiler service."""

import json
import logging
from typing import Annotated, Any, cast

import openpyxl
import openpyxl.styles
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import Field, StringConstraints
from pydantic_settings import BaseSettings

from rts.models import StudyMetadata
from rts.ports.inbound.rev_tran import ReverseTranspilerPort
from rts.ports.outbound.dao import (
    MetadataGridFSDaoPort,
    ResourceNotFoundError,
    WorkbookGridFSDaoPort,
)

log = logging.getLogger(__name__)

__all__ = ["ReverseTranspiler", "SheetNameConfig"]

SheetName = Annotated[str, StringConstraints(max_length=31, min_length=1)]


class SheetNameConfig(BaseSettings):
    """Configuration for sheet names in the spreadsheet output."""

    sheet_names: dict[str, SheetName] = Field(
        ...,
        description="Mapping of worksheet names to display names in the workbook.",
        examples=[
            {
                "analyses": "Analysis",
                "analysis_method_supporting_files": "AnalysisMethodSupportingFile",
            }
        ],
    )


class ReverseTranspiler(ReverseTranspilerPort):
    """Provides functionality to convert study metadata from JSON files back into
    spreadsheet format.
    """

    def __init__(
        self,
        config: SheetNameConfig,
        metadata_dao: MetadataGridFSDaoPort,
        workbook_dao: WorkbookGridFSDaoPort,
    ):
        self._config = config
        self._metadata_dao = metadata_dao
        self._workbook_dao = workbook_dao

    async def upsert_metadata(self, *, study_metadata: StudyMetadata) -> None:
        """Upsert study metadata in the database and reverse transpile it to a workbook.

        If the metadata already exists, it will compare the existing metadata
        with the new one. If they are the same, it will skip the upsert and
        workbook creation. If they differ, it will update the existing metadata
        and create a new workbook, deleting the old one.
        """
        accession = study_metadata.study_accession
        try:
            existing_metadata = await self._metadata_dao.find(id_=accession)
            log.debug(
                "Metadata for accession '%s' already exists, comparing...",
                accession,
            )
            if existing_metadata.model_dump() == study_metadata.model_dump():
                log.debug(
                    "Metadata for accession '%s' has not changed, skipping upsert.",
                    accession,
                )
                return
            else:
                log.info(
                    "Metadata for accession '%s' has changed, updating entry.",
                    accession,
                )
        except ResourceNotFoundError:
            log.debug(
                "No existing metadata found for accession '%s', creating new entry.",
                accession,
            )

        await self._metadata_dao.upsert(data=study_metadata, id_=accession)

        log.debug("Transpiling metadata to workbook for accession '%s'.", accession)
        workbook = self._reverse_transpile(study_metadata)

        log.debug("Workbook created for accession '%s', upserting to DB.", accession)
        await self._workbook_dao.upsert(data=workbook, id_=accession)
        log.debug("Workbook upserted for study accession: %s", accession)

    async def retrieve_metadata(self, *, study_accession: str) -> StudyMetadata:
        """Retrieve study metadata from the DAO by its accession.

        Raises `MetadataNotFoundError` if the metadata does not exist for the
        given study accession.
        """
        try:
            metadata = await self._metadata_dao.find(id_=study_accession)
            return metadata
        except ResourceNotFoundError as err:
            error = self.MetadataNotFoundError(study_accession=study_accession)
            log.error(error)
            raise error from err

    async def delete_metadata(self, *, study_accession: str) -> None:
        """Delete study metadata from the database by its accession.

        This method will always try to delete the associated workbook as well,
        regardless of whether the metadata exists or not.

        Does not raise an error if the metadata or workbook does not exist.
        """
        try:
            await self._metadata_dao.delete(id_=study_accession)
        except ResourceNotFoundError:
            log.debug("Metadata for accession '%s' already deleted.", study_accession)

        await self._workbook_dao.delete(id_=study_accession)
        log.info("Workbook and metadata deleted for accession '%s'.", study_accession)

    async def retrieve_workbook(self, *, study_accession: str) -> bytes:
        """Retrieve the workbook for a given study accession.

        Raises `MetadataNotFoundError` if the workbook does not exist.

        Returns:
        - The workbook as bytes.
        """
        try:
            workbook_data = await self._workbook_dao.find(id_=study_accession)
            log.debug("Workbook found for study accession: %s", study_accession)
            return workbook_data
        except ResourceNotFoundError as err:
            error = self.MetadataNotFoundError(study_accession=study_accession)
            log.error(error)
            raise error from err

    def _translate_sheet_name(self, sheet_name: str) -> str:
        """Rename sheets in the workbook to with configured values.

        If no corresponding value is configured, the name will be returned unchanged
        but truncated to 31 characters if necessary.
        """
        name = self._config.sheet_names.get(sheet_name, sheet_name[:31])
        return name

    def _format_value(self, value: Any) -> Any:
        """Format values for list and dict cells."""
        output = value
        # Handle lists (like types, affiliations, etc.)
        if isinstance(value, list):
            output = ", ".join(str(v) for v in value)
        # Handle dictionaries (like attributes)
        elif isinstance(value, dict):
            output = json.dumps(value)
        return output

    def _reverse_transpile(
        self,
        study_metadata: StudyMetadata,
    ) -> Workbook:
        """Convert StudyMetadata object to a workbook.

        Args:
        - `study_metadata`: The StudyMetadata instance to convert.

        Returns:
        - The metadata as an openpyxl Workbook.
        """
        # Extract the main artifact content
        content = study_metadata.content

        # Create a new workbook
        workbook = openpyxl.Workbook()

        # Remove the default sheet
        default_sheet = cast(Worksheet, workbook.active)
        workbook.remove(default_sheet)

        # Process each property in content (e.g. "analyses", "studies", "samples", etc.)
        for property_name, items in content.items():
            # If there are no items (each item equates to a row), continue to next key
            if not items:
                continue

            # Create a new worksheet for this property
            property_name = self._translate_sheet_name(property_name)
            worksheet: Worksheet = workbook.create_sheet(title=property_name)

            # Get the headers as union of all keys for all items
            # This makes it so we're not reliant on first item having all cols populated
            column_names: set[str] = set(key for row in items for key in row)
            column_headers = list(column_names)

            # Ensure 'alias' is the first column if present
            for idx, special_header in enumerate(["alias", "accession"]):
                if special_header in column_headers:
                    column_headers.remove(special_header)
                    column_headers.insert(idx, special_header)
                else:
                    log.info(  # Unsure of proper log level, or if this log is needed
                        "No '%s' field found in %s property for accession '%s'.",
                        special_header,
                        property_name,
                        study_metadata.study_accession,
                        extra={
                            "study_accession": study_metadata.study_accession,
                            "property": property_name,
                        },
                    )

            # Write the headers to the first row
            for col_idx, header in enumerate(column_headers, 1):
                cell: Cell = worksheet.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = openpyxl.styles.Font(bold=True)
                worksheet.column_dimensions[cell.column_letter].width = 34

            # Process each item and write to the worksheet
            for row_idx, row_data in enumerate(
                items, 2
            ):  # Start from row 2 (after headers)
                for col_idx, header in enumerate(column_headers, 1):
                    cell: Cell = worksheet.cell(row=row_idx, column=col_idx)  # type: ignore
                    value = row_data.get(header)
                    cell.value = self._format_value(value)

        return workbook
